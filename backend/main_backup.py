from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File, Form
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse, StreamingResponse
from jose import JWTError, jwt
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any
import pandas as pd
import os
import json
import hashlib
from pydantic import BaseModel
from dotenv import load_dotenv
import tempfile
from urllib.parse import quote
from sheet_manager import sheet_manager, SheetType
from database import init_db, get_db, Supplier, Product, Order, OrderItem, FileUploadHistory, DailyOrder, WorkDraft, Client
from sqlalchemy.orm import Session
from datetime import date
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()

# Initialize database
init_db()

# Configuration
SECRET_KEY = os.getenv("SECRET_KEY", "your-secret-key-here-change-this-in-production")
ALGORITHM = os.getenv("ALGORITHM", "HS256")
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.getenv("ACCESS_TOKEN_EXPIRE_MINUTES", 1440))
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "gndr_admin")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "gndr12345!")

app = FastAPI(title="GNDR Order Management API", version="1.0.0")

# CORS settings
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:5174"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# OAuth2 scheme
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

# User model
class User(BaseModel):
    username: str
    full_name: Optional[str] = None
    disabled: Optional[bool] = None

class UserInDB(User):
    hashed_password: str

class SheetData(BaseModel):
    sheet_name: str
    data: List[List[Any]]
    columns: List[str]
    rows: int
    cols: int

# Utility functions
def hash_password(password: str) -> str:
    """Simple SHA256 password hashing"""
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(plain_password: str, hashed_password: str) -> bool:
    """Verify password"""
    return hash_password(plain_password) == hashed_password

def get_user(username: str) -> Optional[UserInDB]:
    """Get user by username - simple in-memory user store"""
    if username == ADMIN_USERNAME:
        return UserInDB(
            username=username,
            full_name="GNDR Admin",
            hashed_password=hash_password(ADMIN_PASSWORD),
            disabled=False
        )
    return None

def authenticate_user(username: str, password: str) -> Optional[UserInDB]:
    """Authenticate user"""
    user = get_user(username)
    if not user or not verify_password(password, user.hashed_password):
        return None
    return user

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    """Create JWT access token"""
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=15))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

def serialize_datetime(obj):
    """Convert datetime objects to ISO format strings for JSON serialization"""
    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    raise TypeError(f"Type {type(obj)} not serializable")

def convert_datetime_in_data(data):
    """Recursively convert datetime objects in nested data structures"""
    if isinstance(data, dict):
        return {k: convert_datetime_in_data(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [convert_datetime_in_data(item) for item in data]
    elif isinstance(data, (datetime, date)):
        return data.isoformat()
    else:
        return data

async def get_current_user(token: str = Depends(oauth2_scheme)):
    """Get current user from token"""
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )

    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception

    user = get_user(username)
    if user is None:
        raise credentials_exception
    return user

# Routes
@app.get("/")
async def root():
    """Admin panel with server control"""
    html_content = """
    <!DOCTYPE html>
    <html>
    <head>
        <title>GNDR 서버 관리</title>
        <meta charset="utf-8">
        <style>
            body {
                font-family: Arial, sans-serif;
                max-width: 800px;
                margin: 50px auto;
                padding: 20px;
                background-color: #f5f5f5;
            }
            .container {
                background: white;
                padding: 30px;
                border-radius: 10px;
                box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            }
            h1 {
                color: #333;
                margin-bottom: 30px;
            }
            .status {
                padding: 15px;
                background: #e8f5e9;
                border-left: 4px solid #4caf50;
                margin-bottom: 20px;
                border-radius: 4px;
            }
            .button-group {
                display: flex;
                gap: 10px;
                margin-top: 20px;
            }
            button {
                padding: 12px 24px;
                font-size: 16px;
                border: none;
                border-radius: 5px;
                cursor: pointer;
                transition: all 0.3s;
            }
            .restart-btn {
                background: #ff9800;
                color: white;
            }
            .restart-btn:hover {
                background: #f57c00;
            }
            .info-btn {
                background: #2196f3;
                color: white;
            }
            .info-btn:hover {
                background: #1976d2;
            }
            .message {
                padding: 15px;
                margin-top: 20px;
                border-radius: 4px;
                display: none;
            }
            .success {
                background: #d4edda;
                border: 1px solid #c3e6cb;
                color: #155724;
            }
            .error {
                background: #f8d7da;
                border: 1px solid #f5c6cb;
                color: #721c24;
            }
            .info {
                background: #d1ecf1;
                border: 1px solid #bee5eb;
                color: #0c5460;
            }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>🚀 GNDR 주문 관리 시스템</h1>

            <div class="status">
                <strong>서버 상태:</strong> 정상 작동 중 ✅
            </div>

            <div class="button-group">
                <button class="restart-btn" onclick="restartServer()">🔄 서버 재시작</button>
                <button class="info-btn" onclick="showInfo()">ℹ️ 서버 정보</button>
            </div>

            <div class="button-group" style="margin-top: 10px;">
                <button style="background: #4caf50; color: white;" onclick="fixPColumn()">🔧 P열 수정 (기존 데이터)</button>
            </div>

            <hr style="margin: 30px 0; border: none; border-top: 2px solid #e0e0e0;">

            <h2 style="color: #333; margin-bottom: 20px;">📊 데이터 관리</h2>

            <div class="status" id="statsBox" style="background: #e3f2fd; border-left: 4px solid #2196f3;">
                <strong>현재 데이터:</strong>
                <div id="stats" style="margin-top: 10px; font-size: 14px;">
                    로딩 중...
                </div>
            </div>

            <div class="button-group">
                <button style="background: #2196f3; color: white;" onclick="refreshStats()">🔄 통계 새로고침</button>
            </div>

            <div class="button-group" style="margin-top: 10px;">
                <button style="background: #f44336; color: white;" onclick="clearPayments()">🗑️ 입금 내역 전체 삭제</button>
                <button style="background: #ff9800; color: white;" onclick="clearFiles()">🗑️ 파일 내역 전체 삭제</button>
                <button style="background: #9c27b0; color: white;" onclick="clearOrders()">🗑️ 발주 내역 전체 삭제</button>
            </div>

            <div class="button-group" style="margin-top: 10px;">
                <button style="background: #d32f2f; color: white; font-weight: bold;" onclick="clearAll()">⚠️ 전체 데이터 삭제</button>
            </div>

            <div id="message" class="message"></div>
        </div>

        <script>
            async function restartServer() {
                const messageDiv = document.getElementById('message');
                messageDiv.textContent = '서버를 재시작하고 있습니다...';
                messageDiv.className = 'message info';
                messageDiv.style.display = 'block';

                try {
                    const response = await fetch('/admin/restart', {
                        method: 'POST'
                    });
                    const data = await response.json();

                    if (response.ok) {
                        messageDiv.textContent = '✅ ' + data.message;
                        messageDiv.className = 'message success';

                        setTimeout(() => {
                            messageDiv.textContent = '서버 재시작 완료! 페이지를 새로고침합니다...';
                            setTimeout(() => location.reload(), 2000);
                        }, 2000);
                    } else {
                        messageDiv.textContent = '❌ 재시작 실패: ' + data.detail;
                        messageDiv.className = 'message error';
                    }
                } catch (error) {
                    messageDiv.textContent = '❌ 오류 발생: ' + error.message;
                    messageDiv.className = 'message error';
                }
            }

            function showInfo() {
                const messageDiv = document.getElementById('message');
                messageDiv.innerHTML = `
                    <strong>서버 정보:</strong><br>
                    - API URL: http://localhost:8000<br>
                    - Frontend URL: http://localhost:3000<br>
                    - 문서: http://localhost:8000/docs<br>
                    - 상태: 정상 작동 중
                `;
                messageDiv.className = 'message info';
                messageDiv.style.display = 'block';
            }

            async function fixPColumn() {
                const messageDiv = document.getElementById('message');
                messageDiv.textContent = 'P열을 수정하고 있습니다...';
                messageDiv.className = 'message info';
                messageDiv.style.display = 'block';

                try {
                    const response = await fetch('/admin/fix-p-column', {
                        method: 'POST'
                    });
                    const data = await response.json();

                    if (response.ok) {
                        messageDiv.textContent = '✅ ' + data.message;
                        messageDiv.className = 'message success';
                    } else {
                        messageDiv.textContent = '❌ 수정 실패: ' + data.detail;
                        messageDiv.className = 'message error';
                    }
                } catch (error) {
                    messageDiv.textContent = '❌ 오류 발생: ' + error.message;
                    messageDiv.className = 'message error';
                }
            }

            // 데이터 관리 함수들
            let authToken = null;

            async function getAuthToken() {
                if (authToken) return authToken;

                try {
                    const response = await fetch('/token', {
                        method: 'POST',
                        headers: { 'Content-Type': 'application/x-www-form-urlencoded' },
                        body: 'username=gndr_admin&password=gndr1234!!'
                    });

                    if (response.ok) {
                        const data = await response.json();
                        authToken = data.access_token;
                        return authToken;
                    }
                } catch (error) {
                    console.error('인증 실패:', error);
                }
                return null;
            }

            async function refreshStats() {
                const statsDiv = document.getElementById('stats');
                statsDiv.innerHTML = '로딩 중...';

                try {
                    const token = await getAuthToken();
                    if (!token) {
                        statsDiv.innerHTML = '<span style="color: #f44336;">❌ 인증 실패</span>';
                        return;
                    }

                    const response = await fetch('/admin/stats', {
                        headers: { 'Authorization': `Bearer ${token}` }
                    });

                    if (response.ok) {
                        const data = await response.json();
                        const stats = data.stats;
                        statsDiv.innerHTML = `
                            • 입금 내역: <strong>${stats.payments}개</strong><br>
                            • 파일 내역: <strong>${stats.files}개</strong><br>
                            • 발주 내역: <strong>${stats.orders}개</strong><br>
                            • 임시 저장: <strong>${stats.drafts}개</strong>
                        `;
                    } else {
                        statsDiv.innerHTML = '<span style="color: #f44336;">❌ 통계 조회 실패</span>';
                    }
                } catch (error) {
                    statsDiv.innerHTML = '<span style="color: #f44336;">❌ 오류: ' + error.message + '</span>';
                }
            }

            async function clearPayments() {
                if (!confirm('⚠️ 모든 입금 내역을 삭제하시겠습니까?\\n\\n이 작업은 되돌릴 수 없습니다!')) {
                    return;
                }

                const messageDiv = document.getElementById('message');
                messageDiv.textContent = '입금 내역을 삭제하고 있습니다...';
                messageDiv.className = 'message info';
                messageDiv.style.display = 'block';

                try {
                    const token = await getAuthToken();
                    if (!token) {
                        messageDiv.textContent = '❌ 인증 실패';
                        messageDiv.className = 'message error';
                        return;
                    }

                    const response = await fetch('/admin/payments/clear-all', {
                        method: 'DELETE',
                        headers: { 'Authorization': `Bearer ${token}` }
                    });

                    const data = await response.json();

                    if (response.ok) {
                        messageDiv.textContent = '✅ ' + data.message;
                        messageDiv.className = 'message success';
                        await refreshStats();
                    } else {
                        messageDiv.textContent = '❌ 삭제 실패: ' + data.detail;
                        messageDiv.className = 'message error';
                    }
                } catch (error) {
                    messageDiv.textContent = '❌ 오류 발생: ' + error.message;
                    messageDiv.className = 'message error';
                }
            }

            async function clearFiles() {
                if (!confirm('⚠️ 모든 파일 내역을 삭제하시겠습니까?\\n\\n이 작업은 되돌릴 수 없습니다!')) {
                    return;
                }

                const messageDiv = document.getElementById('message');
                messageDiv.textContent = '파일 내역을 삭제하고 있습니다...';
                messageDiv.className = 'message info';
                messageDiv.style.display = 'block';

                try {
                    const token = await getAuthToken();
                    if (!token) {
                        messageDiv.textContent = '❌ 인증 실패';
                        messageDiv.className = 'message error';
                        return;
                    }

                    const response = await fetch('/admin/files/clear-all', {
                        method: 'DELETE',
                        headers: { 'Authorization': `Bearer ${token}` }
                    });

                    const data = await response.json();

                    if (response.ok) {
                        messageDiv.textContent = '✅ ' + data.message;
                        messageDiv.className = 'message success';
                        await refreshStats();
                    } else {
                        messageDiv.textContent = '❌ 삭제 실패: ' + data.detail;
                        messageDiv.className = 'message error';
                    }
                } catch (error) {
                    messageDiv.textContent = '❌ 오류 발생: ' + error.message;
                    messageDiv.className = 'message error';
                }
            }

            async function clearOrders() {
                if (!confirm('⚠️ 모든 발주 내역을 삭제하시겠습니까?\\n\\n이 작업은 되돌릴 수 없습니다!')) {
                    return;
                }

                const messageDiv = document.getElementById('message');
                messageDiv.textContent = '발주 내역을 삭제하고 있습니다...';
                messageDiv.className = 'message info';
                messageDiv.style.display = 'block';

                try {
                    const token = await getAuthToken();
                    if (!token) {
                        messageDiv.textContent = '❌ 인증 실패';
                        messageDiv.className = 'message error';
                        return;
                    }

                    const response = await fetch('/admin/orders/clear-all', {
                        method: 'DELETE',
                        headers: { 'Authorization': `Bearer ${token}` }
                    });

                    const data = await response.json();

                    if (response.ok) {
                        messageDiv.textContent = '✅ ' + data.message;
                        messageDiv.className = 'message success';
                        await refreshStats();
                    } else {
                        messageDiv.textContent = '❌ 삭제 실패: ' + data.detail;
                        messageDiv.className = 'message error';
                    }
                } catch (error) {
                    messageDiv.textContent = '❌ 오류 발생: ' + error.message;
                    messageDiv.className = 'message error';
                }
            }

            async function clearAll() {
                if (!confirm('⚠️⚠️⚠️ 경고! ⚠️⚠️⚠️\\n\\n입금, 파일, 발주 내역을 모두 삭제하시겠습니까?\\n\\n이 작업은 되돌릴 수 없습니다!')) {
                    return;
                }

                if (!confirm('정말로 모든 데이터를 삭제하시겠습니까?\\n\\n마지막 확인입니다!')) {
                    return;
                }

                const messageDiv = document.getElementById('message');
                messageDiv.textContent = '모든 데이터를 삭제하고 있습니다...';
                messageDiv.className = 'message info';
                messageDiv.style.display = 'block';

                try {
                    const token = await getAuthToken();
                    if (!token) {
                        messageDiv.textContent = '❌ 인증 실패';
                        messageDiv.className = 'message error';
                        return;
                    }

                    // 입금 내역 삭제
                    await fetch('/admin/payments/clear-all', {
                        method: 'DELETE',
                        headers: { 'Authorization': `Bearer ${token}` }
                    });

                    // 파일 내역 삭제
                    await fetch('/admin/files/clear-all', {
                        method: 'DELETE',
                        headers: { 'Authorization': `Bearer ${token}` }
                    });

                    // 발주 내역 삭제
                    await fetch('/admin/orders/clear-all', {
                        method: 'DELETE',
                        headers: { 'Authorization': `Bearer ${token}` }
                    });

                    messageDiv.textContent = '✅ 모든 데이터가 삭제되었습니다';
                    messageDiv.className = 'message success';
                    await refreshStats();
                } catch (error) {
                    messageDiv.textContent = '❌ 오류 발생: ' + error.message;
                    messageDiv.className = 'message error';
                }
            }

            // 페이지 로드 시 통계 새로고침
            window.addEventListener('load', refreshStats);
        </script>
    </body>
    </html>
    """
    from fastapi.responses import HTMLResponse
    return HTMLResponse(content=html_content)

@app.post("/token")
async def login(form_data: OAuth2PasswordRequestForm = Depends()):
    """Login endpoint"""
    user = authenticate_user(form_data.username, form_data.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )

    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.username}, expires_delta=access_token_expires
    )
    return {"access_token": access_token, "token_type": "bearer"}

@app.post("/admin/restart")
async def restart_server():
    """Restart the server by touching main.py to trigger auto-reload"""
    try:
        import os
        import time

        # Touch main.py to trigger uvicorn auto-reload
        main_file = os.path.abspath(__file__)
        os.utime(main_file, None)

        logger.info("Server restart triggered")
        return {"success": True, "message": "서버 재시작 중... 잠시 후 자동으로 재연결됩니다."}
    except Exception as e:
        logger.error(f"Server restart failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"재시작 실패: {str(e)}")

@app.post("/admin/fix-p-column")
async def fix_p_column(db: Session = Depends(get_db)):
    """기존 저장된 모든 주문서의 P열을 재계산하여 수정"""
    try:
        # 모든 일별 주문서 가져오기
        orders = db.query(DailyOrder).all()
        fixed_count = 0

        for order in orders:
            if not order.data:
                continue

            # P열 재계산
            processed_data = []
            for row_idx, row in enumerate(order.data):
                if row_idx <= 2:  # 헤더 행은 그대로
                    processed_data.append(row)
                else:
                    # 데이터 행 - P열 처리
                    row_copy = list(row)

                    if len(row_copy) > 15:  # P열이 존재하면 (인덱스 15)
                        try:
                            # L, M, N열 값 가져오기 (인덱스 11, 12, 13)
                            l_val = row_copy[11] if len(row_copy) > 11 else 0
                            m_val = row_copy[12] if len(row_copy) > 12 else 0
                            n_val = row_copy[13] if len(row_copy) > 13 else 0
                            o_val = row_copy[14] if len(row_copy) > 14 else 0

                            # 숫자로 변환
                            l_num = float(l_val) if l_val not in [None, "", " "] else 0
                            m_num = float(m_val) if m_val not in [None, "", " "] else 0
                            n_num = float(n_val) if n_val not in [None, "", " "] else 0
                            o_num = float(o_val) if o_val not in [None, "", " "] else 0

                            lmn_sum = l_num + m_num + n_num

                            # P열 값 설정
                            if lmn_sum != o_num:
                                row_copy[15] = "차이 있음"
                            else:
                                row_copy[15] = ""
                        except (ValueError, TypeError):
                            row_copy[15] = ""

                    processed_data.append(row_copy)

            # 데이터베이스 업데이트
            order.data = processed_data
            order.updated_at = datetime.now()
            fixed_count += 1

        db.commit()
        logger.info(f"Fixed P column for {fixed_count} orders")
        return {"success": True, "message": f"{fixed_count}개의 주문서 P열이 수정되었습니다.", "fixed_count": fixed_count}

    except Exception as e:
        db.rollback()
        logger.error(f"Failed to fix P column: {str(e)}")
        raise HTTPException(status_code=500, detail=f"P열 수정 실패: {str(e)}")

@app.get("/users/me")
async def get_current_user_info(current_user: User = Depends(get_current_user)):
    """Get current user info"""
    return current_user

# Save order data to database
def save_order_data_to_db(sheets: List[Dict], file_path: str, db: Session):
    """Save order data to database from sheets"""
    try:
        for sheet in sheets:
            # Create or update Order
            order = Order(
                order_date=date.today(),
                order_type=sheet.get("sheet_type", "주문서"),
                sheet_name=sheet.get("sheet_name", "Unknown"),
                total_amount=0
            )
            db.add(order)
            db.flush()  # Get the order ID

            # Process each row of data
            for idx, row in enumerate(sheet.get("data", [])[1:]):  # Skip header row
                if idx == 0:  # Skip the first data row (column names)
                    continue

                # Extract supplier info (assuming column structure)
                supplier_name = row[0] if row else None
                if not supplier_name:
                    continue

                # Check if supplier exists or create new one
                supplier = db.query(Supplier).filter_by(name=supplier_name).first()
                if not supplier:
                    supplier = Supplier(
                        name=supplier_name,
                        address=row[1] if len(row) > 1 else None,
                        phone=row[2] if len(row) > 2 else None,
                        mobile=row[3] if len(row) > 3 else None
                    )
                    db.add(supplier)
                    db.flush()

                # Product info
                product_code = row[4] if len(row) > 4 else None
                product_name = row[5] if len(row) > 5 else None

                if product_code or product_name:
                    # Check if product exists or create new one
                    product = db.query(Product).filter_by(code=product_code).first()
                    if not product:
                        product = Product(
                            code=product_code,
                            name=product_name,
                            supplier_name=row[5] if len(row) > 5 else None,
                            supplier_option=row[6] if len(row) > 6 else None,
                            price=float(row[7]) if len(row) > 7 and row[7] else 0
                        )
                        db.add(product)
                        db.flush()

                    # Create OrderItem
                    order_item = OrderItem(
                        order_id=order.id,
                        product_id=product.id,
                        new_order_qty=int(row[8]) if len(row) > 8 and row[8] else 0,
                        undelivered_qty=int(row[9]) if len(row) > 9 and row[9] else 0,
                        exchange_qty=int(row[10]) if len(row) > 10 and row[10] else 0,
                        janggi_qty=int(row[11]) if len(row) > 11 and row[11] else 0,
                        janggi_undelivered=int(row[12]) if len(row) > 12 and row[12] else 0,
                        janggi_exchange=int(row[13]) if len(row) > 13 and row[13] else 0,
                        received_qty=int(row[14]) if len(row) > 14 and row[14] else 0,
                        difference_qty=int(row[15]) if len(row) > 15 and row[15] else 0,
                        uncle_comment=row[16] if len(row) > 16 else None,
                        gndr_comment=row[17] if len(row) > 17 else None
                    )
                    db.add(order_item)

            # Update order's supplier_id with the first supplier found
            if supplier:
                order.supplier_id = supplier.id

        # Save file upload history
        file_history = FileUploadHistory(
            filename=os.path.basename(file_path),
            file_path=file_path,
            sheet_count=len(sheets),
            row_count=sum(len(sheet.get("data", [])) for sheet in sheets),
            status="success",
            user="system"
        )
        db.add(file_history)

        db.commit()
        return True

    except Exception as e:
        db.rollback()
        print(f"Error saving to database: {e}")
        return False

# Excel handling routes
@app.get("/excel/check")
async def check_excel(
    current_user: User = Depends(get_current_user)
):
    """저장된 엑셀 파일이 있는지 확인"""
    try:
        has_file = sheet_manager.has_loaded_file()
        filename = None

        if has_file:
            file_info = sheet_manager.get_current_file_info()
            if file_info and 'file_path' in file_info:
                import os
                filename = os.path.basename(file_info['file_path'])

        return JSONResponse(content={
            "success": True,
            "has_file": has_file,
            "filename": filename
        })
    except Exception as e:
        logger.error(f"Error checking excel: {str(e)}")
        return JSONResponse(
            status_code=500,
            content={"success": False, "error": str(e)}
        )

@app.get("/excel/load")
async def load_excel_file(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Load the default Excel file with work state if available"""
    try:
        # Default file path - Updated to correct file
        file_path = "/Users/pablokim/gndr-orma/docs/references/0828가나다란 주문서.xlsx"

        # Load using sheet manager
        result = sheet_manager.load_excel_file(file_path)

        if not result["success"]:
            raise HTTPException(status_code=500, detail=result.get("error", "Failed to load Excel file"))

        # Check for saved work state
        draft = db.query(WorkDraft).filter(
            WorkDraft.user == current_user.username,
            WorkDraft.draft_type == "order_work",
            WorkDraft.expires_at > datetime.now()
        ).order_by(WorkDraft.updated_at.desc()).first()

        if draft:
            # Restore work state
            result["work_state"] = {
                "is_order_receipt_uploaded": bool(draft.is_order_receipt_uploaded),
                "is_receipt_slip_uploaded": bool(draft.is_receipt_slip_uploaded),
                "row_colors": draft.row_colors if draft.row_colors else {},
                "row_text_colors": draft.row_text_colors if draft.row_text_colors else {},
                "checked_rows": draft.checked_rows if draft.checked_rows else {},
                "duplicate_products": draft.duplicate_products if draft.duplicate_products else {},
                "sheets_data": draft.sheets_data if draft.sheets_data else None
            }

            # If sheets_data exists in draft, use it instead of default file
            if draft.sheets_data:
                result["sheets"] = draft.sheets_data

            logger.info(f"Restored work state for user {current_user.username}")

        return result

    except Exception as e:
        logger.error(f"Error loading excel file: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/excel/upload-order-receipt")
async def upload_order_receipt_file(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """Upload order receipt file for validation and merging"""
    try:
        # Check file extension
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Only Excel files are allowed")

        # Save uploaded file temporarily
        suffix = '.xls' if file.filename.endswith('.xls') else '.xlsx'
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp_file:
            content = await file.read()
            tmp_file.write(content)
            tmp_path = tmp_file.name

        # Read the order receipt file with original filename
        result = sheet_manager.load_excel_file(tmp_path, original_filename=file.filename)

        if not result["success"]:
            os.unlink(tmp_path)
            raise HTTPException(status_code=400, detail="Failed to read order receipt file")

        # Get the first sheet data
        receipt_data = result["sheets"][0]["data"]

        # Clean up temp file
        os.unlink(tmp_path)

        # Convert datetime objects in data to ISO format strings
        receipt_data = convert_datetime_in_data(receipt_data)

        # Return the receipt data for merging
        return JSONResponse(content={
            "success": True,
            "data": receipt_data,
            "message": "주문 입고 파일이 검증되었습니다."
        })

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error processing order receipt: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/excel/upload-receipt-slip")
async def upload_receipt_slip_file(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """Upload receipt slip file (입고전표) for matching product codes and store names"""
    try:
        # Check file extension
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Only Excel files are allowed")

        # Save uploaded file temporarily
        suffix = '.xls' if file.filename.endswith('.xls') else '.xlsx'
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp_file:
            content = await file.read()
            tmp_file.write(content)
            tmp_path = tmp_file.name

        # Read the receipt slip file with original filename
        result = sheet_manager.load_excel_file(tmp_path, original_filename=file.filename)

        if not result["success"]:
            os.unlink(tmp_path)
            raise HTTPException(status_code=400, detail="Failed to read receipt slip file")

        # Get the first sheet data
        receipt_slip_data = result["sheets"][0]["data"]

        # Clean up temp file
        os.unlink(tmp_path)

        # Return the receipt slip data for processing
        return JSONResponse(content={
            "success": True,
            "data": receipt_slip_data,
            "message": "입고전표가 성공적으로 업로드되었습니다."
        })

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error processing receipt slip: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/excel/upload")
async def upload_excel_file(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Upload a new Excel file"""
    try:
        # Check file extension
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Only Excel files are allowed")

        # Save uploaded file temporarily
        suffix = '.xls' if file.filename.endswith('.xls') else '.xlsx'
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp_file:
            content = await file.read()
            tmp_file.write(content)
            tmp_path = tmp_file.name

        # Use sheet manager to load file with original filename
        result = sheet_manager.load_excel_file(tmp_path, original_filename=file.filename)

        if not result["success"]:
            # Clean up temporary file on error
            os.unlink(tmp_path)
            raise HTTPException(status_code=500, detail=result.get("error", "Unknown error"))

        # Save to database
        if result.get("sheets"):
            save_order_data_to_db(result["sheets"], tmp_path, db)

        # Clean up temporary file
        os.unlink(tmp_path)

        result["filename"] = file.filename
        return result

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/excel/update")
async def update_excel_data(
    sheet_data: SheetData,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update Excel data and save to database"""
    try:
        # Update the sheet in the sheet manager's cache
        sheet_manager.update_sheet_data(sheet_data.sheet_name, sheet_data.data)

        # Also update the database with the new order data
        # Clear existing order items for this sheet and re-save
        # This is a simplified approach - in production you'd want more sophisticated update logic
        sheets_to_save = [{
            "sheet_name": sheet_data.sheet_name,
            "data": sheet_data.data,
            "columns": sheet_data.columns,
            "rows": sheet_data.rows,
            "cols": sheet_data.cols
        }]

        # Save to database
        save_order_data_to_db(sheets_to_save, f"Updated-{sheet_data.sheet_name}", db)

        return {
            "success": True,
            "message": "데이터가 성공적으로 저장되었습니다",
            "sheet_name": sheet_data.sheet_name,
            "rows_updated": sheet_data.rows
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/excel/cached")
async def get_cached_sheets(current_user: User = Depends(get_current_user)):
    """Get list of cached sheets"""
    try:
        cached = sheet_manager.get_cached_sheets()
        return {
            "success": True,
            "cached_sheets": cached
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class ExportData(BaseModel):
    data: List[List[Any]]
    columns: List[str]
    file_name: str
    sheet_name: str = "Sheet1"

class SaveDailyOrderData(BaseModel):
    date: date  # 날짜
    order_type: str  # 'order', 'receipt', 'voucher'
    sheet_name: str
    data: List[List[Any]]
    columns: List[str]
    notes: Optional[str] = None

class WorkDraftData(BaseModel):
    """작업 중간 저장 데이터"""
    draft_type: str  # 'spreadsheet', 'order', 'receipt'
    sheets_data: List[Dict[str, Any]]
    selected_sheet: int = 0
    row_colors: Dict[int, str] = {}
    row_text_colors: Dict[int, str] = {}
    duplicate_products: Dict[int, str] = {}
    is_order_receipt_uploaded: bool = False
    is_receipt_slip_uploaded: bool = False
    checked_rows: Dict[int, bool] = {}
    hide_checked: bool = False
    description: Optional[str] = None
    session_id: Optional[str] = None

@app.post("/excel/export")
async def export_excel(
    export_data: ExportData,
    current_user: User = Depends(get_current_user)
):
    """Export data to Excel file"""
    try:
        output_file = sheet_manager.export_to_excel(
            data=export_data.data,
            columns=export_data.columns,
            file_name=export_data.file_name,
            sheet_name=export_data.sheet_name
        )

        # Return file as download
        return FileResponse(
            path=output_file,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename=os.path.basename(output_file)
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/orders/statistics")
async def get_order_statistics(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get order statistics and summary"""
    try:
        # Get statistics
        total_suppliers = db.query(Supplier).count()
        total_products = db.query(Product).count()
        total_orders = db.query(Order).count()
        total_order_items = db.query(OrderItem).count()

        # Calculate total order value
        total_order_value = db.query(Order).with_entities(
            db.func.sum(Order.total_amount)
        ).scalar() or 0

        # Get recent orders
        recent_orders = db.query(Order).order_by(
            Order.created_at.desc()
        ).limit(10).all()

        # Get items with undelivered quantities
        undelivered_items = db.query(OrderItem).filter(
            OrderItem.undelivered_qty > 0
        ).count()

        return {
            "success": True,
            "summary": {
                "total_suppliers": total_suppliers,
                "total_products": total_products,
                "total_orders": total_orders,
                "total_order_items": total_order_items,
                "total_order_value": float(total_order_value),
                "undelivered_items": undelivered_items
            },
            "recent_orders": [
                {
                    "id": order.id,
                    "order_date": order.order_date.isoformat() if order.order_date else None,
                    "order_type": order.order_type,
                    "sheet_name": order.sheet_name,
                    "total_amount": float(order.total_amount) if order.total_amount else 0,
                    "items_count": len(order.items)
                }
                for order in recent_orders
            ]
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Daily Order Management Endpoints
@app.post("/daily-orders/save")
async def save_daily_order(
    order_data: SaveDailyOrderData,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """일별 주문서 저장"""
    try:
        # Process P column (index 15) for difference check
        # P열을 "차이 있음" 또는 빈칸으로 변경
        processed_data = []
        for row_idx, row in enumerate(order_data.data):
            if row_idx <= 2:  # Keep header rows as is
                processed_data.append(row)
            else:
                # Data row - process P column
                row_copy = list(row)  # Make a copy

                if len(row_copy) > 15:  # P column exists (index 15)
                    # Get L, M, N values (indices 11, 12, 13)
                    l_val = row_copy[11] if len(row_copy) > 11 and row_copy[11] else 0
                    m_val = row_copy[12] if len(row_copy) > 12 and row_copy[12] else 0
                    n_val = row_copy[13] if len(row_copy) > 13 and row_copy[13] else 0

                    # Get O value (index 14)
                    o_val = row_copy[14] if len(row_copy) > 14 and row_copy[14] else 0

                    try:
                        # Convert to numbers
                        l_num = float(l_val) if l_val else 0
                        m_num = float(m_val) if m_val else 0
                        n_num = float(n_val) if n_val else 0
                        o_num = float(o_val) if o_val else 0

                        lmn_sum = l_num + m_num + n_num

                        # Set P column value
                        if lmn_sum != o_num:
                            row_copy[15] = "차이 있음"
                        else:
                            row_copy[15] = ""
                    except (ValueError, TypeError):
                        # If conversion fails, keep original value
                        pass

                processed_data.append(row_copy)

        # Check if order already exists
        existing = db.query(DailyOrder).filter(
            DailyOrder.date == order_data.date,
            DailyOrder.order_type == order_data.order_type,
            DailyOrder.sheet_name == order_data.sheet_name
        ).first()

        if existing:
            # Update existing order
            existing.data = processed_data
            existing.columns = order_data.columns
            existing.notes = order_data.notes
            existing.updated_at = datetime.now()

            # Calculate metadata
            existing.total_items = len(processed_data) - 4  # Exclude header rows
            db.commit()

            return {
                "success": True,
                "message": "주문서가 업데이트되었습니다",
                "order_id": existing.id
            }
        else:
            # Create new order
            new_order = DailyOrder(
                date=order_data.date,
                order_type=order_data.order_type,
                sheet_name=order_data.sheet_name,
                data=processed_data,
                columns=order_data.columns,
                notes=order_data.notes,
                created_by=current_user.username,
                total_items=len(processed_data) - 4  # Exclude header rows
            )

            db.add(new_order)
            db.commit()
            db.refresh(new_order)

            return {
                "success": True,
                "message": "주문서가 저장되었습니다",
                "order_id": new_order.id
            }

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/daily-orders/list")
async def get_daily_orders(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    order_type: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """일별 주문서 목록 조회"""
    try:
        query = db.query(DailyOrder)

        if start_date:
            query = query.filter(DailyOrder.date >= start_date)
        if end_date:
            query = query.filter(DailyOrder.date <= end_date)
        if order_type:
            query = query.filter(DailyOrder.order_type == order_type)

        orders = query.order_by(DailyOrder.date.desc()).all()

        return {
            "success": True,
            "orders": [
                {
                    "id": order.id,
                    "date": order.date.isoformat(),
                    "order_type": order.order_type,
                    "sheet_name": order.sheet_name,
                    "total_items": order.total_items or 0,
                    "total_quantity": order.total_quantity or 0,
                    "total_amount": order.total_amount or 0,
                    "created_at": order.created_at.isoformat(),
                    "created_by": order.created_by,
                    "notes": order.notes
                }
                for order in orders
            ]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/daily-orders/{order_id}")
async def get_daily_order(
    order_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """특정 일별 주문서 조회"""
    try:
        order = db.query(DailyOrder).filter(DailyOrder.id == order_id).first()

        if not order:
            raise HTTPException(status_code=404, detail="주문서를 찾을 수 없습니다")

        return {
            "success": True,
            "order": {
                "id": order.id,
                "date": order.date.isoformat(),
                "order_type": order.order_type,
                "sheet_name": order.sheet_name,
                "data": order.data,
                "columns": order.columns,
                "total_items": order.total_items,
                "created_at": order.created_at.isoformat(),
                "created_by": order.created_by,
                "notes": order.notes
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/daily-orders/{order_id}")
async def delete_daily_order(
    order_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """일별 주문서 삭제"""
    try:
        order = db.query(DailyOrder).filter(DailyOrder.id == order_id).first()
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")

        db.delete(order)
        db.commit()

        return {"success": True, "message": "Order deleted successfully"}
    except Exception as e:
        logger.error(f"Error deleting daily order {order_id}: {str(e)}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/daily-orders/{order_id}/download")
async def download_daily_order(
    order_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """일별 주문서를 엑셀 파일로 다운로드"""
    try:
        order = db.query(DailyOrder).filter(DailyOrder.id == order_id).first()
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")

        # 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = order.sheet_name

        # 스타일 정의
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        orange_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        blue_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        red_font = Font(color="FF0000", bold=True)

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 데이터 쓰기
        data = order.data
        if data:
            for row_idx, row_data in enumerate(data, start=1):
                for col_idx, value in enumerate(row_data, start=1):
                    # P열(16번 컬럼)은 무조건 재계산
                    if col_idx == 16 and row_idx > 2:  # 헤더 이후 행에서만 처리
                        # L, M, N열 값 가져오기 (인덱스 11, 12, 13)
                        try:
                            l_val = row_data[11] if len(row_data) > 11 else 0
                            m_val = row_data[12] if len(row_data) > 12 else 0
                            n_val = row_data[13] if len(row_data) > 13 else 0
                            o_val = row_data[14] if len(row_data) > 14 else 0

                            # 숫자로 변환 (빈 값이나 None은 0으로)
                            l_num = float(l_val) if l_val not in [None, "", " "] else 0
                            m_num = float(m_val) if m_val not in [None, "", " "] else 0
                            n_num = float(n_val) if n_val not in [None, "", " "] else 0
                            o_num = float(o_val) if o_val not in [None, "", " "] else 0

                            lmn_sum = l_num + m_num + n_num

                            # 합계와 O열 값이 다르면 "차이 있음", 같으면 빈칸
                            if lmn_sum != o_num:
                                cell = ws.cell(row=row_idx, column=col_idx, value="차이 있음")
                                cell.font = red_font
                            else:
                                cell = ws.cell(row=row_idx, column=col_idx, value="")
                        except (ValueError, TypeError, IndexError) as e:
                            # 변환 실패 시 빈칸
                            logger.warning(f"P열 계산 실패 row={row_idx}: {e}")
                            cell = ws.cell(row=row_idx, column=col_idx, value="")
                    else:
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)

                    # 테두리 추가
                    cell.border = thin_border

                    # 헤더 행 스타일
                    if row_idx <= 2:
                        cell.fill = header_fill
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    # 데이터 행 색상 적용
                    elif row_idx > 2:
                        # 색상 적용 (L, M, N열: 12, 13, 14)
                        if 12 <= col_idx <= 14:
                            cell.fill = orange_fill
                        # O, P열: 15, 16
                        elif 15 <= col_idx <= 16:
                            cell.fill = blue_fill
                        # Q~W열: 17~23
                        elif 17 <= col_idx <= 23:
                            cell.fill = green_fill

        # 열 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # 메모리에 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 파일명 생성
        order_type_name = {
            'order': '발주서',
            'receipt': '주문입고',
            'voucher': '입고전표'
        }.get(order.order_type, order.order_type)

        filename = f"{order.date}_{order_type_name}_{order.sheet_name}.xlsx"
        encoded_filename = quote(filename)

        return StreamingResponse(
            io.BytesIO(output.getvalue()),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            }
        )
    except Exception as e:
        logger.error(f"Error downloading daily order {order_id}: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# ==========================================
# 작업 중간 저장 (Work Draft) API
# ==========================================

@app.post("/work-drafts/save")
async def save_work_draft(
    draft_data: WorkDraftData,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """작업 중간 저장"""
    try:
        # 기존 임시 저장 데이터가 있는지 확인
        existing_draft = db.query(WorkDraft).filter(
            WorkDraft.user == current_user.username,
            WorkDraft.draft_type == draft_data.draft_type
        ).first()

        # 만료 시간 설정 (24시간 후)
        expires_at = datetime.now() + timedelta(hours=24)

        if existing_draft:
            # 업데이트
            existing_draft.sheets_data = draft_data.sheets_data
            existing_draft.selected_sheet = draft_data.selected_sheet
            existing_draft.row_colors = draft_data.row_colors
            existing_draft.row_text_colors = draft_data.row_text_colors
            existing_draft.duplicate_products = draft_data.duplicate_products
            existing_draft.is_order_receipt_uploaded = 1 if draft_data.is_order_receipt_uploaded else 0
            existing_draft.is_receipt_slip_uploaded = 1 if draft_data.is_receipt_slip_uploaded else 0
            existing_draft.checked_rows = draft_data.checked_rows
            existing_draft.hide_checked = 1 if draft_data.hide_checked else 0
            existing_draft.description = draft_data.description
            existing_draft.session_id = draft_data.session_id
            existing_draft.updated_at = datetime.now()
            existing_draft.expires_at = expires_at

            db.commit()

            return {
                "success": True,
                "message": "작업이 중간 저장되었습니다",
                "draft_id": existing_draft.id,
                "updated_at": existing_draft.updated_at.isoformat()
            }
        else:
            # 새로 생성
            new_draft = WorkDraft(
                user=current_user.username,
                draft_type=draft_data.draft_type,
                sheets_data=draft_data.sheets_data,
                selected_sheet=draft_data.selected_sheet,
                row_colors=draft_data.row_colors,
                row_text_colors=draft_data.row_text_colors,
                duplicate_products=draft_data.duplicate_products,
                is_order_receipt_uploaded=1 if draft_data.is_order_receipt_uploaded else 0,
                is_receipt_slip_uploaded=1 if draft_data.is_receipt_slip_uploaded else 0,
                checked_rows=draft_data.checked_rows,
                hide_checked=1 if draft_data.hide_checked else 0,
                description=draft_data.description,
                session_id=draft_data.session_id,
                expires_at=expires_at
            )

            db.add(new_draft)
            db.commit()
            db.refresh(new_draft)

            return {
                "success": True,
                "message": "작업이 중간 저장되었습니다",
                "draft_id": new_draft.id,
                "created_at": new_draft.created_at.isoformat()
            }

    except Exception as e:
        db.rollback()
        logger.error(f"Error saving work draft: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/work-drafts/load")
async def load_work_draft(
    draft_type: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """작업 중간 저장 데이터 불러오기"""
    try:
        # 만료되지 않은 가장 최근 데이터 찾기
        draft = db.query(WorkDraft).filter(
            WorkDraft.user == current_user.username,
            WorkDraft.draft_type == draft_type,
            WorkDraft.expires_at > datetime.now()
        ).order_by(WorkDraft.updated_at.desc()).first()

        if not draft:
            return {
                "success": True,
                "has_draft": False,
                "message": "저장된 작업이 없습니다"
            }

        return {
            "success": True,
            "has_draft": True,
            "draft": {
                "id": draft.id,
                "draft_type": draft.draft_type,
                "sheets_data": draft.sheets_data,
                "selected_sheet": draft.selected_sheet,
                "row_colors": draft.row_colors,
                "row_text_colors": draft.row_text_colors,
                "duplicate_products": draft.duplicate_products,
                "is_order_receipt_uploaded": bool(draft.is_order_receipt_uploaded),
                "is_receipt_slip_uploaded": bool(draft.is_receipt_slip_uploaded),
                "checked_rows": draft.checked_rows if draft.checked_rows else {},
                "hide_checked": bool(draft.hide_checked),
                "description": draft.description,
                "created_at": draft.created_at.isoformat(),
                "updated_at": draft.updated_at.isoformat(),
                "expires_at": draft.expires_at.isoformat()
            }
        }

    except Exception as e:
        logger.error(f"Error loading work draft: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/work-drafts/delete")
async def delete_work_draft(
    draft_type: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """작업 중간 저장 데이터 삭제"""
    try:
        draft = db.query(WorkDraft).filter(
            WorkDraft.user == current_user.username,
            WorkDraft.draft_type == draft_type
        ).first()

        if draft:
            db.delete(draft)
            db.commit()
            return {
                "success": True,
                "message": "임시 저장 데이터가 삭제되었습니다"
            }
        else:
            return {
                "success": True,
                "message": "삭제할 데이터가 없습니다"
            }

    except Exception as e:
        db.rollback()
        logger.error(f"Error deleting work draft: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/work-drafts/cleanup-expired")
async def cleanup_expired_drafts(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """만료된 임시 저장 데이터 정리"""
    try:
        deleted_count = db.query(WorkDraft).filter(
            WorkDraft.expires_at < datetime.now()
        ).delete()

        db.commit()

        return {
            "success": True,
            "message": f"{deleted_count}개의 만료된 임시 저장 데이터가 삭제되었습니다",
            "deleted_count": deleted_count
        }

    except Exception as e:
        db.rollback()
        logger.error(f"Error cleaning up expired drafts: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# ==================== Payment Management Endpoints ====================

class PaymentDataRequest(BaseModel):
    payment_date: str
    data: List[List[Any]]
    columns: List[str]
    created_by: Optional[str] = None

@app.post("/payments/save")
async def save_payment_data(
    request: PaymentDataRequest,
    db: Session = Depends(get_db)
):
    """입금 내역 저장 - 중복 제거 후 체크된 항목을 일자별로 누적 저장"""
    try:
        from datetime import datetime
        from database import PaymentRecord

        payment_date = datetime.strptime(request.payment_date, "%Y-%m-%d").date()
        saved_count = 0
        skipped_count = 0

        # 헤더 4행 추출 (인덱스 0-3)
        header_rows = request.data[:4] if len(request.data) >= 4 else []

        # 헤더를 sheet_manager에 저장하여 나중에 불러올 수 있도록 함
        if header_rows and sheet_manager.loaded_sheets:
            for sheet_key, sheet_data in sheet_manager.loaded_sheets.items():
                if 'data' in sheet_data:
                    # 기존 시트 데이터의 헤더 4행을 현재 헤더로 업데이트
                    sheet_data['data'][:4] = header_rows
                    break

        # 해당 날짜의 기존 입금 내역 조회
        existing_payments = db.query(PaymentRecord).filter(
            PaymentRecord.payment_date == payment_date
        ).all()

        # 헤더 저장 여부 확인 - 첫 번째 레코드에 헤더 정보를 저장
        has_header_record = False
        for payment in existing_payments:
            if isinstance(payment.original_data, dict) and payment.original_data.get('_is_header'):
                has_header_record = True
                # 기존 헤더 레코드 업데이트
                payment.original_data['header_rows'] = header_rows
                break

        # 헤더 레코드가 없으면 생성
        if not has_header_record and header_rows:
            header_record = PaymentRecord(
                payment_date=payment_date,
                company_name='_HEADER_',
                product_code='_HEADER_',
                product_name='헤더 정보',
                product_option='',
                unit_price=0.0,
                receipt_qty=0,
                payment_amount=0.0,
                original_data={'_is_header': True, 'header_rows': header_rows},
                created_by=request.created_by
            )
            db.add(header_record)

        # 중복 체크를 위한 키 생성 함수 (거래처명 + 상품코드 + 상품명 + 옵션 + 입고량)
        def create_payment_key(company, code, name, option, qty):
            return f"{company}|{code or ''}|{name or ''}|{option or ''}|{qty}"

        # 기존 입금 내역의 키 세트 생성 (헤더 레코드 제외)
        existing_keys = set()
        for payment in existing_payments:
            # 헤더 레코드는 중복 체크에서 제외
            if payment.company_name == '_HEADER_':
                continue

            key = create_payment_key(
                payment.company_name,
                payment.product_code,
                payment.product_name,
                payment.product_option,
                payment.receipt_qty
            )
            existing_keys.add(key)

        # 데이터 행만 처리 (인덱스 4부터)
        for i in range(4, len(request.data)):
            row = request.data[i]

            # 거래처명이 없으면 건너뛰기
            if not row[0]:
                continue

            # 새 입금 내역 키 생성
            company_name = str(row[0])
            product_code = str(row[4]) if len(row) > 4 and row[4] else None
            product_name = str(row[5]) if len(row) > 5 and row[5] else None
            product_option = str(row[6]) if len(row) > 6 and row[6] else None
            receipt_qty = int(row[14]) if len(row) > 14 and row[14] else 0

            new_key = create_payment_key(
                company_name,
                product_code,
                product_name,
                product_option,
                receipt_qty
            )

            # 중복 체크 - 이미 존재하면 건너뛰기
            if new_key in existing_keys:
                skipped_count += 1
                continue

            # PaymentRecord 생성
            payment_record = PaymentRecord(
                payment_date=payment_date,
                company_name=company_name,
                product_code=product_code,
                product_name=product_name,
                product_option=product_option,
                unit_price=float(row[7]) if len(row) > 7 and row[7] else 0.0,  # H열: 원가
                receipt_qty=receipt_qty,
                payment_amount=float(row[19]) if len(row) > 19 and row[19] else 0.0,  # T열: 입금액
                original_data=row,  # 전체 행 데이터 저장
                created_by=request.created_by
            )

            db.add(payment_record)
            existing_keys.add(new_key)  # 새로 추가한 항목도 키 세트에 추가
            saved_count += 1

        db.commit()

        message = f"{saved_count}개 항목이 {request.payment_date} 입금 내역으로 저장되었습니다"
        if skipped_count > 0:
            message += f" ({skipped_count}개 중복 항목 제외)"

        return {
            "success": True,
            "message": message,
            "saved_count": saved_count,
            "skipped_count": skipped_count,
            "payment_date": request.payment_date
        }

    except Exception as e:
        db.rollback()
        logger.error(f"Error saving payment data: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/payments/delete")
async def delete_payment_records(
    payment_ids: List[int],
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """선택된 입금 내역 삭제"""
    try:
        from database import PaymentRecord

        deleted_count = 0
        for payment_id in payment_ids:
            payment = db.query(PaymentRecord).filter(PaymentRecord.id == payment_id).first()
            if payment:
                db.delete(payment)
                deleted_count += 1

        db.commit()

        return {
            "success": True,
            "message": f"{deleted_count}개 항목이 삭제되었습니다",
            "deleted_count": deleted_count
        }

    except Exception as e:
        db.rollback()
        logger.error(f"Error deleting payment records: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/payments/delete-all")
async def delete_all_payment_records(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """모든 입금 내역 삭제"""
    try:
        from database import PaymentRecord

        deleted_count = db.query(PaymentRecord).delete()
        db.commit()

        return {
            "success": True,
            "message": f"총 {deleted_count}개의 입금 내역이 삭제되었습니다",
            "deleted_count": deleted_count
        }

    except Exception as e:
        db.rollback()
        logger.error(f"Error deleting all payment records: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/payments/date/{payment_date}")
async def get_payments_by_date(
    payment_date: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """특정 날짜의 입금 내역 조회 - 현재 로드된 시트의 헤더와 함께 반환"""
    try:
        from datetime import datetime
        from database import PaymentRecord

        date_obj = datetime.strptime(payment_date, "%Y-%m-%d").date()

        payments = db.query(PaymentRecord).filter(
            PaymentRecord.payment_date == date_obj
        ).order_by(PaymentRecord.company_name, PaymentRecord.id).all()

        # 업체별 합계 계산 (헤더 레코드 제외)
        company_totals = {}
        for payment in payments:
            # 헤더 레코드는 제외
            if isinstance(payment.original_data, dict) and payment.original_data.get('_is_header'):
                continue
            if payment.company_name not in company_totals:
                company_totals[payment.company_name] = 0
            company_totals[payment.company_name] += payment.payment_amount

        # 헤더 4행 가져오기
        header_rows = []

        # 1. DB에 저장된 헤더 레코드에서 가져오기 (우선순위)
        for payment in payments:
            if isinstance(payment.original_data, dict) and payment.original_data.get('_is_header'):
                header_rows = payment.original_data.get('header_rows', [])
                break

        # 2. 헤더가 없으면 sheet_manager에서 가져오기
        if not header_rows and sheet_manager.loaded_sheets:
            first_sheet = list(sheet_manager.loaded_sheets.values())[0]
            if first_sheet and 'data' in first_sheet:
                sheet_data = first_sheet['data']
                if len(sheet_data) >= 4:
                    header_rows = sheet_data[:4]

        # 헤더 레코드 제외하고 실제 데이터만 반환
        actual_payments = [p for p in payments if not (isinstance(p.original_data, dict) and p.original_data.get('_is_header'))]

        return {
            "success": True,
            "payment_date": payment_date,
            "total_count": len(actual_payments),
            "total_amount": sum(p.payment_amount for p in actual_payments),
            "company_totals": company_totals,
            "header_rows": header_rows,  # 헤더 4행 추가
            "payments": [
                {
                    "id": p.id,
                    "company_name": p.company_name,
                    "product_code": p.product_code,
                    "product_name": p.product_name,
                    "product_option": p.product_option,
                    "unit_price": p.unit_price,
                    "receipt_qty": p.receipt_qty,
                    "payment_amount": p.payment_amount,
                    "original_data": p.original_data,
                    "created_at": p.created_at.isoformat() if p.created_at else None
                }
                for p in actual_payments
            ]
        }

    except Exception as e:
        logger.error(f"Error fetching payments by date: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/payments/range")
async def get_payments_by_range(
    start: str,
    end: str,
    db: Session = Depends(get_db)
):
    """기간별 입금 내역 조회"""
    try:
        from datetime import datetime
        from database import PaymentRecord

        start_date = datetime.strptime(start, "%Y-%m-%d").date()
        end_date = datetime.strptime(end, "%Y-%m-%d").date()

        payments = db.query(PaymentRecord).filter(
            PaymentRecord.payment_date >= start_date,
            PaymentRecord.payment_date <= end_date
        ).order_by(PaymentRecord.payment_date, PaymentRecord.company_name).all()

        # 일자별, 업체별 합계 계산
        date_totals = {}
        company_totals = {}

        for payment in payments:
            date_key = payment.payment_date.isoformat()
            if date_key not in date_totals:
                date_totals[date_key] = 0
            date_totals[date_key] += payment.payment_amount

            if payment.company_name not in company_totals:
                company_totals[payment.company_name] = 0
            company_totals[payment.company_name] += payment.payment_amount

        return {
            "success": True,
            "start_date": start,
            "end_date": end,
            "total_count": len(payments),
            "total_amount": sum(p.payment_amount for p in payments),
            "date_totals": date_totals,
            "company_totals": company_totals,
            "payments": [
                {
                    "id": p.id,
                    "payment_date": p.payment_date.isoformat(),
                    "company_name": p.company_name,
                    "product_code": p.product_code,
                    "product_name": p.product_name,
                    "product_option": p.product_option,
                    "unit_price": p.unit_price,
                    "receipt_qty": p.receipt_qty,
                    "payment_amount": p.payment_amount,
                    "created_at": p.created_at.isoformat() if p.created_at else None
                }
                for p in payments
            ]
        }

    except Exception as e:
        logger.error(f"Error fetching payments by range: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================================
# Order Management APIs (발주 관리)
# ============================================

class OrderRequest(BaseModel):
    order_date: str  # YYYY-MM-DD 형식
    company_name: str
    order_type: str  # '교환', '미송', '기타'
    items: list  # 발주 항목 리스트

@app.post("/orders/save")
async def save_order_records(request: OrderRequest, db: Session = Depends(get_db)):
    """발주 내역 저장 (교환/미송 등)"""
    try:
        from database import OrderRecord
        from datetime import datetime

        order_date_obj = datetime.strptime(request.order_date, "%Y-%m-%d").date()
        saved_count = 0

        for item in request.items:
            # 거래처명 (A열, index 0)
            company_name = item[0] if len(item) > 0 else request.company_name
            # 상품코드 (E열, index 4)
            product_code = item[4] if len(item) > 4 else None
            # 공급처상품명 (F열, index 5)
            product_name = item[5] if len(item) > 5 else None
            # 공급처옵션 (G열, index 6)
            product_option = item[6] if len(item) > 6 else None
            # 원가 (H열, index 7)
            unit_price = float(item[7]) if len(item) > 7 and item[7] else 0.0
            # 발주수량 - 실제로는 입고량 컬럼 사용 (O열, index 14)
            order_qty = int(item[14]) if len(item) > 14 and item[14] else 0
            # 발주금액 계산
            order_amount = unit_price * order_qty

            order_record = OrderRecord(
                order_date=order_date_obj,
                company_name=company_name,
                order_type=request.order_type,
                product_code=product_code,
                product_name=product_name,
                product_option=product_option,
                unit_price=unit_price,
                order_qty=order_qty,
                order_amount=order_amount,
                original_data=item,
                created_by=request.company_name  # 또는 별도 사용자 정보
            )

            db.add(order_record)
            saved_count += 1

        db.commit()

        return {
            "success": True,
            "message": f"{saved_count}건의 발주 내역이 저장되었습니다",
            "order_date": request.order_date,
            "order_type": request.order_type
        }

    except Exception as e:
        logger.error(f"Error saving order records: {str(e)}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/orders/list")
async def list_order_records(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """발주 내역 목록 조회 (날짜별 그룹화)"""
    try:
        from database import OrderRecord
        from sqlalchemy import func

        # 날짜별로 그룹화
        dates = db.query(OrderRecord.order_date).distinct().order_by(OrderRecord.order_date.desc()).all()

        result = []
        for (order_date,) in dates:
            # 해당 날짜의 발주 내역 조회
            orders = db.query(OrderRecord).filter(
                OrderRecord.order_date == order_date
            ).all()

            # 발주 유형별로 분류
            orders_by_type = {}
            for order in orders:
                if order.order_type not in orders_by_type:
                    orders_by_type[order.order_type] = []
                orders_by_type[order.order_type].append({
                    "id": order.id,
                    "company_name": order.company_name,
                    "product_code": order.product_code,
                    "product_name": order.product_name,
                    "product_option": order.product_option,
                    "unit_price": order.unit_price,
                    "order_qty": order.order_qty,
                    "order_amount": order.order_amount,
                    "status": order.status,
                    "original_data": order.original_data,
                    "created_at": order.created_at.isoformat() if order.created_at else None
                })

            result.append({
                "order_date": order_date.isoformat(),
                "orders_by_type": orders_by_type,
                "total_count": len(orders)
            })

        return {
            "success": True,
            "data": result
        }

    except Exception as e:
        logger.error(f"Error listing order records: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/orders/date/{date}")
async def get_orders_by_date(
    date: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """특정 날짜의 발주 내역 조회 - 현재 로드된 시트의 헤더와 함께 반환"""
    try:
        from database import OrderRecord
        from datetime import datetime

        order_date = datetime.strptime(date, "%Y-%m-%d").date()
        orders = db.query(OrderRecord).filter(
            OrderRecord.order_date == order_date
        ).all()

        # 헤더 4행 가져오기
        header_rows = []

        # 1. DB에 저장된 헤더 레코드에서 가져오기 (우선순위)
        for order in orders:
            if isinstance(order.original_data, dict) and order.original_data.get('_is_header'):
                header_rows = order.original_data.get('header_rows', [])
                break

        # 2. 헤더가 없으면 sheet_manager에서 가져오기
        if not header_rows and sheet_manager.loaded_sheets:
            first_sheet = list(sheet_manager.loaded_sheets.values())[0]
            if first_sheet and 'data' in first_sheet:
                sheet_data = first_sheet['data']
                if len(sheet_data) >= 4:
                    header_rows = sheet_data[:4]

        # 헤더 레코드 제외하고 실제 데이터만 반환
        actual_orders = [o for o in orders if not (isinstance(o.original_data, dict) and o.original_data.get('_is_header'))]

        return {
            "success": True,
            "order_date": date,
            "header_rows": header_rows,  # 헤더 4행 추가
            "orders": [
                {
                    "id": o.id,
                    "company_name": o.company_name,
                    "order_type": o.order_type,
                    "product_code": o.product_code,
                    "product_name": o.product_name,
                    "product_option": o.product_option,
                    "unit_price": o.unit_price,
                    "order_qty": o.order_qty,
                    "order_amount": o.order_amount,
                    "status": o.status,
                    "original_data": o.original_data,
                    "created_at": o.created_at.isoformat() if o.created_at else None
                }
                for o in actual_orders
            ]
        }

    except Exception as e:
        logger.error(f"Error fetching orders by date: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================================
# Saved Files Management APIs
# ============================================

class SaveFilesRequest(BaseModel):
    date: str  # MMDD 형식
    matched_data: dict  # 매칭 파일 데이터
    normal_data: dict  # 정상 파일 데이터
    error_data: dict  # 오류 파일 데이터
    created_by: str

@app.post("/files/save-three-files")
async def save_three_files(request: SaveFilesRequest, db: Session = Depends(get_db)):
    """입금관리로 보낼 때 3개 파일을 자동 저장"""
    try:
        from database import SavedFile
        import os

        # 저장 디렉토리 생성
        save_dir = "./saved_files"
        os.makedirs(save_dir, exist_ok=True)

        files_to_save = [
            ("matched", f"{request.date}주문입고-매칭.xlsx", request.matched_data),
            ("normal", f"{request.date}주문입고-정상.xlsx", request.normal_data),
            ("error", f"{request.date}주문입고-오류.xlsx", request.error_data)
        ]

        saved_files = []

        for file_type, file_name, file_data in files_to_save:
            file_path = os.path.join(save_dir, file_name)

            # 기존 파일이 있으면 삭제
            existing = db.query(SavedFile).filter(
                SavedFile.date == request.date,
                SavedFile.file_type == file_type
            ).first()

            if existing:
                if os.path.exists(existing.file_path):
                    os.remove(existing.file_path)
                db.delete(existing)
                db.commit()

            # 엑셀 파일 생성
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font

            wb = Workbook()
            ws = wb.active

            data = file_data.get('data', [])
            row_colors = file_data.get('row_colors', {})
            row_text_colors = file_data.get('row_text_colors', {})

            # 데이터 쓰기
            for row_idx, row_data in enumerate(data):
                for col_idx, cell_value in enumerate(row_data):
                    cell = ws.cell(row=row_idx+1, column=col_idx+1, value=cell_value)

                    # 행 색상 적용
                    if str(row_idx) in row_colors:
                        color = row_colors[str(row_idx)].replace('#', '')
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

                    # 텍스트 색상 적용
                    if str(row_idx) in row_text_colors:
                        color = row_text_colors[str(row_idx)].replace('#', '')
                        cell.font = Font(color=color)

            wb.save(file_path)

            # DB에 저장
            new_file = SavedFile(
                date=request.date,
                file_type=file_type,
                file_name=file_name,
                file_path=file_path,
                sheet_data=file_data.get('data', []),
                columns=file_data.get('columns', []),
                row_colors=row_colors,
                row_text_colors=row_text_colors,
                total_rows=len(file_data.get('data', [])),
                created_by=request.created_by
            )

            db.add(new_file)
            db.commit()
            db.refresh(new_file)

            saved_files.append({
                "file_type": file_type,
                "file_name": file_name,
                "file_path": file_path,
                "total_rows": new_file.total_rows
            })

        return {
            "success": True,
            "message": "3개 파일이 모두 저장되었습니다",
            "files": saved_files
        }

    except Exception as e:
        logger.error(f"Error saving three files: {str(e)}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/files/list")
async def list_saved_files(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    """저장된 파일 목록 조회 (날짜별로 그룹화)"""
    try:
        from database import SavedFile
        from sqlalchemy import func

        # 날짜별로 그룹화
        dates = db.query(SavedFile.date).distinct().order_by(SavedFile.date.desc()).all()

        result = []
        for (date_str,) in dates:
            files = db.query(SavedFile).filter(SavedFile.date == date_str).all()

            file_dict = {}
            for f in files:
                file_dict[f.file_type] = {
                    "id": f.id,
                    "file_name": f.file_name,
                    "file_path": f.file_path,
                    "total_rows": f.total_rows,
                    "created_at": f.created_at.isoformat() if f.created_at else None
                }

            result.append({
                "date": date_str,
                "files": file_dict
            })

        return {
            "success": True,
            "data": result
        }

    except Exception as e:
        logger.error(f"Error listing saved files: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/files/view/{file_id}")
async def view_saved_file(file_id: int, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    """저장된 파일 데이터 조회"""
    try:
        from database import SavedFile

        file = db.query(SavedFile).filter(SavedFile.id == file_id).first()

        if not file:
            raise HTTPException(status_code=404, detail="파일을 찾을 수 없습니다")

        return {
            "success": True,
            "data": {
                "file_name": file.file_name,
                "date": file.date,
                "file_type": file.file_type,
                "sheet_data": file.sheet_data,
                "columns": file.columns,
                "row_colors": file.row_colors,
                "row_text_colors": file.row_text_colors,
                "total_rows": file.total_rows,
                "created_at": file.created_at.isoformat() if file.created_at else None
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error viewing saved file: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/files/download/{file_id}")
async def download_saved_file(file_id: int, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    """저장된 파일 다운로드"""
    try:
        from database import SavedFile

        file = db.query(SavedFile).filter(SavedFile.id == file_id).first()

        if not file:
            raise HTTPException(status_code=404, detail="파일을 찾을 수 없습니다")

        if not os.path.exists(file.file_path):
            raise HTTPException(status_code=404, detail="파일이 존재하지 않습니다")

        return FileResponse(
            path=file.file_path,
            filename=file.file_name,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error downloading saved file: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================================
# Admin Management APIs (관리자 전용)
# ============================================

@app.delete("/admin/payments/clear-all")
async def admin_clear_all_payments(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """[관리자] 모든 입금 관리 내역 삭제"""
    try:
        from database import PaymentRecord

        deleted_count = db.query(PaymentRecord).delete()
        db.commit()

        logger.info(f"Admin cleared all payment records: {deleted_count} records deleted by {current_user.username}")

        return {
            "success": True,
            "message": f"총 {deleted_count}개의 입금 내역이 삭제되었습니다",
            "deleted_count": deleted_count
        }

    except Exception as e:
        db.rollback()
        logger.error(f"Error clearing all payments: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/admin/files/clear-all")
async def admin_clear_all_files(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """[관리자] 모든 파일 관리 내역 삭제"""
    try:
        from database import SavedFile
        import os

        # DB에서 파일 레코드 조회
        files = db.query(SavedFile).all()
        deleted_db_count = 0
        deleted_file_count = 0

        # 물리적 파일 삭제
        for file in files:
            if file.file_path and os.path.exists(file.file_path):
                try:
                    os.remove(file.file_path)
                    deleted_file_count += 1
                except Exception as e:
                    logger.warning(f"Failed to delete file {file.file_path}: {e}")

        # DB 레코드 삭제
        deleted_db_count = db.query(SavedFile).delete()
        db.commit()

        logger.info(f"Admin cleared all files: {deleted_db_count} DB records, {deleted_file_count} files deleted by {current_user.username}")

        return {
            "success": True,
            "message": f"총 {deleted_db_count}개의 파일 내역이 삭제되었습니다 (물리적 파일 {deleted_file_count}개 삭제)",
            "deleted_db_records": deleted_db_count,
            "deleted_files": deleted_file_count
        }

    except Exception as e:
        db.rollback()
        logger.error(f"Error clearing all files: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/admin/orders/clear-all")
async def admin_clear_all_orders(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """[관리자] 모든 발주 관리 내역 삭제"""
    try:
        from database import OrderRecord

        deleted_count = db.query(OrderRecord).delete()
        db.commit()

        logger.info(f"Admin cleared all order records: {deleted_count} records deleted by {current_user.username}")

        return {
            "success": True,
            "message": f"총 {deleted_count}개의 발주 내역이 삭제되었습니다",
            "deleted_count": deleted_count
        }

    except Exception as e:
        db.rollback()
        logger.error(f"Error clearing all orders: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/admin/stats")
async def admin_get_stats(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """[관리자] 시스템 통계 조회"""
    try:
        from database import PaymentRecord, SavedFile, OrderRecord, WorkDraft

        payment_count = db.query(PaymentRecord).count()
        file_count = db.query(SavedFile).count()
        order_count = db.query(OrderRecord).count()
        draft_count = db.query(WorkDraft).count()

        return {
            "success": True,
            "stats": {
                "payments": payment_count,
                "files": file_count,
                "orders": order_count,
                "drafts": draft_count
            }
        }

    except Exception as e:
        logger.error(f"Error getting admin stats: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))